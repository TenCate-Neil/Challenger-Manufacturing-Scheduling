#!/usr/bin/env python3
"""
Tests for the per-item batch requirements module.

The core (max colour widths, the yarn_lbs join, bobbin counts) is
dependency-free; the extractor block test needs openpyxl and skips cleanly
when it is not importable.

Runs with pytest, or standalone with no dependencies:

    python test_item_requirements.py
"""

import item_requirements as ir

try:
    import openpyxl
    from extract_turf_layout import extract_yarn_lbs
    _HAVE_OPENPYXL = True
except Exception as _exc:  # noqa: BLE001
    _HAVE_OPENPYXL = False
    _IMPORT_ERROR = _exc


def _roll(*segments, panels=None):
    roll = {"segments": [{"color_code": c, "width_in": w}
                         for c, w in segments]}
    if panels is not None:
        roll["additional_panel_layouts"] = [
            {"segments": [{"color_code": c, "width_in": w} for c, w in panel]}
            for panel in panels
        ]
    return roll


def _yarn(position, yarn_type, *colors):
    return {
        "yarn_position": position,
        "yarn_type": yarn_type,
        "colors": [{"color_code": code, "color_name": code,
                    "sku": sku, "lbs_needed": lbs}
                   for code, sku, lbs in colors],
    }


def _extraction(yarn_lbs, rolls):
    return {"source_file": "S.xlsx", "yarn_lbs": yarn_lbs, "rolls": rolls}


# --- max colour widths ------------------------------------------------------
def test_same_colour_segments_sum_within_one_roll():
    # 5" + 5" of WHI in one roll counts as 10" — it beats a single 8" run in
    # another roll, so the max is 10 and bobbins are 30.
    rolls = [_roll(("FG", 172), ("WHI", 5), ("FG", 100), ("WHI", 5)),
             _roll(("FG", 174), ("WHI", 8))]
    widths = ir.max_color_widths(rolls)
    assert widths["WHI"] == 10
    ext = _extraction([_yarn("Y1", "T", ("WHI", 111, 42.0))], rolls)
    items = ir.item_requirements(ext)
    assert items[0]["max_width_in"] == 10
    assert items[0]["bobbins_required"] == 30


def test_max_is_taken_across_rolls():
    # The planners' worked example: colour widths of 5, 3, and 8 inches in
    # three different rolls -> the widest single-roll requirement is 8", so
    # 24 bobbins.
    rolls = [_roll(("FG", 177), ("WHI", 5)),
             _roll(("FG", 179), ("WHI", 3)),
             _roll(("FG", 174), ("WHI", 8))]
    ext = _extraction([_yarn("Y1", "T", ("WHI", 111, 42.0))], rolls)
    items = ir.item_requirements(ext)
    assert items[0]["max_width_in"] == 8
    assert items[0]["bobbins_required"] == 24


def test_fractional_widths_round_bobbins_up():
    # 5.5" x 3 = 16.5 -> 17 bobbins; a fraction of a bobbin cannot be dressed.
    rolls = [_roll(("FG", 176.5), ("WHI", 5.5))]
    ext = _extraction([_yarn("Y1", "T", ("WHI", 111, 42.0))], rolls)
    items = ir.item_requirements(ext)
    assert items[0]["max_width_in"] == 5.5
    assert items[0]["bobbins_required"] == 17


def test_additional_panel_layouts_are_independent_candidates():
    # A panel layout is cut from the same threading: its per-colour sum is a
    # candidate of its own, never added to the roll's own segments. Here the
    # panel's 12" WHI wins over the roll's 5" — and does NOT become 17".
    rolls = [_roll(("FG", 177), ("WHI", 5),
                   panels=[[("FG", 170), ("WHI", 12)]])]
    widths = ir.max_color_widths(rolls)
    assert widths["WHI"] == 12
    ext = _extraction([_yarn("Y1", "T", ("WHI", 111, 42.0))], rolls)
    items = ir.item_requirements(ext)
    assert items[0]["bobbins_required"] == 36


# --- join warnings ----------------------------------------------------------
def test_block_colour_absent_from_rolls_gets_zero_bobbins_and_warning():
    rolls = [_roll(("FG", 182))]
    ext = _extraction([_yarn("Y1", "T", ("FG", 111, 42.0),
                             ("BLK", 222, 3.0))], rolls)
    warnings = []
    items = ir.item_requirements(ext, warnings)
    blk = [i for i in items if i["color_code"] == "BLK"][0]
    assert blk["max_width_in"] == 0
    assert blk["bobbins_required"] == 0
    assert any("BLK" in w and "never appears in any roll" in w
               for w in warnings)


def test_roll_colour_missing_from_block_warns_once():
    # WHI is tufted in two rolls but priced nowhere in the yarn_lbs block:
    # exactly one warning, not one per roll.
    rolls = [_roll(("FG", 177), ("WHI", 5)),
             _roll(("FG", 174), ("WHI", 8))]
    ext = _extraction([_yarn("Y1", "T", ("FG", 111, 42.0))], rolls)
    warnings = []
    ir.item_requirements(ext, warnings)
    missing = [w for w in warnings if "WHI" in w and "missing" in w]
    assert len(missing) == 1


def test_missing_yarn_lbs_key_returns_empty_with_warning():
    warnings = []
    items = ir.item_requirements({"rolls": [_roll(("FG", 182))]}, warnings)
    assert items == []
    assert any("yarn_lbs" in w for w in warnings)


# --- ordering and shape -----------------------------------------------------
def test_items_follow_block_order():
    rolls = [_roll(("FG", 100), ("WHI", 50), ("BLK", 32))]
    ext = _extraction([_yarn("Y1", "A", ("FG", 1, 10.0), ("WHI", 2, 5.0)),
                       _yarn("Y2", "B", ("FG", 3, 12.0), ("BLK", 4, 2.0))],
                      rolls)
    items = ir.item_requirements(ext)
    assert [(i["yarn_position"], i["color_code"], i["item_number"])
            for i in items] == [("Y1", "FG", 1), ("Y1", "WHI", 2),
                                ("Y2", "FG", 3), ("Y2", "BLK", 4)]
    assert all(i["yarn_type"] in ("A", "B") for i in items)
    assert items[0]["lbs_needed"] == 10.0


# --- extractor block --------------------------------------------------------
def _build_yarn_lbs_worksheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B638"] = "Yarn SKUs & Total Lbs. Needed"
    ws["B639"] = "Yarn Type"
    # Colour headers on row 639 (E..Z range).
    ws["E639"], ws["F639"], ws["G639"] = "FG", "WHI", "BLK"
    # Y1: FG has a real SKU; WHI is '#N/A' but needs lbs (-> sku None +
    # warning); BLK has a SKU but 0 lbs (still included).
    ws["A640"], ws["B640"], ws["D640"] = "Y1", "Type A", "SKU"
    ws["E640"], ws["F640"], ws["G640"] = 111, "#N/A", 222
    ws["B641"] = "Yarn Lbs."
    ws["E641"], ws["F641"], ws["G641"] = 100.5, 50.0, 0
    # Y2: shared string SKU must be kept as-is, unused colour stays absent.
    ws["A642"], ws["B642"], ws["D642"] = "Y2", "Type B", "SKU"
    ws["E642"], ws["F642"], ws["G642"] = "145190A", "145190A", "#N/A"
    ws["B643"] = "Yarn Lbs."
    ws["E643"], ws["F643"], ws["G643"] = 12.5, 3.25, 0
    # Y3 and Y4 unused (yarn type 0 / empty) -> skipped entirely.
    ws["A644"], ws["B644"], ws["D644"] = "Y3", 0, "SKU"
    ws["B645"] = "Yarn Lbs."
    ws["A646"], ws["D646"] = "Y4", "SKU"
    ws["B647"] = "Yarn Lbs."
    return ws


def test_extract_yarn_lbs_from_worksheet():
    if not _HAVE_OPENPYXL:
        return  # skip: openpyxl unavailable
    warnings = []
    result = extract_yarn_lbs(_build_yarn_lbs_worksheet(), legend={},
                              warnings=warnings)
    assert [y["yarn_position"] for y in result] == ["Y1", "Y2"]
    assert result[0]["yarn_type"] == "Type A"

    y1 = {c["color_code"]: c for c in result[0]["colors"]}
    assert set(y1) == {"FG", "WHI", "BLK"}
    assert y1["FG"]["sku"] == 111 and y1["FG"]["lbs_needed"] == 100.5
    # lbs > 0 with '#N/A' SKU: included with sku None, plus a warning.
    assert y1["WHI"]["sku"] is None and y1["WHI"]["lbs_needed"] == 50.0
    assert any("Y1" in w and "WHI" in w and "no SKU" in w for w in warnings)
    # SKU present with 0 lbs: still an entry.
    assert y1["BLK"]["sku"] == 222 and y1["BLK"]["lbs_needed"] == 0

    y2 = {c["color_code"]: c for c in result[1]["colors"]}
    assert set(y2) == {"FG", "WHI"}  # '#N/A' + 0 lbs colour dropped
    assert y2["FG"]["sku"] == "145190A"  # string SKUs are not coerced
    # No anchor warnings — the only warning is the missing-SKU one.
    assert [w for w in warnings if "Anchor mismatch" in w] == []


def test_extract_yarn_lbs_warns_on_moved_anchor():
    if not _HAVE_OPENPYXL:
        return  # skip: openpyxl unavailable
    ws = _build_yarn_lbs_worksheet()
    ws["B638"] = "Something Else"
    warnings = []
    extract_yarn_lbs(ws, legend={}, warnings=warnings)
    assert any("yarn lbs header" in w for w in warnings)


def _run_standalone():
    tests = [v for name, v in sorted(globals().items())
             if name.startswith("test_") and callable(v)]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"  PASS  {test.__name__}")
        except AssertionError as exc:
            failures += 1
            print(f"  FAIL  {test.__name__}: {exc}")
        except Exception as exc:  # noqa: BLE001
            failures += 1
            print(f"  ERROR {test.__name__}: {exc!r}")
    print(f"\n{len(tests) - failures}/{len(tests)} passed")
    return 1 if failures else 0


if __name__ == "__main__":
    import sys
    sys.exit(_run_standalone())
