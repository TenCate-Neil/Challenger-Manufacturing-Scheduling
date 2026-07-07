#!/usr/bin/env python3
"""
Extracts manufacturing data from turf "FIELD LAYOUT" Excel workbooks.

Usage:
    python extract_turf_layout.py FILE.xlsx [FILE2.xlsx ...] [-o out_dir]

For each input workbook, writes a JSON file with:
  - general_information   (project/customer/PO/etc, cells B2:K10)
  - product_specifications (yarn types, gauge, pile height, weights, cells M2:U12,
                             plus the Brand field which is resolved from an embedded
                             logo image rather than cell text - see resolve_brand_logo)
  - mfg_summary            (roll/weight/truck totals, cells M15:T16)
  - yarn_skus              (creel position -> yarn type -> available colors + SKUs,
                             rows 671-676)
  - rolls                  (the roll-by-roll cut layout, row 684 onward, grouped by
                             setup change / creel change)
  - warnings               (anything that didn't match the expected template shape)
"""

import argparse
import hashlib
import json
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

SHEET_NAME = "FIELD LAYOUT"
DROPBOX_SHEET_NAME = "Drop Box sheet"

# sha256 of the embedded brand-logo PNGs, seen identically across every sample
# workbook (they're shared corporate assets baked into the template's hidden
# "PRODUCT SPECS " sheet). Extend this map if a new brand logo shows up.
KNOWN_BRAND_LOGOS = {
    "58d70c5952580fe47b7d9e600fce652c88251addcb5670ced92518dae483cde6": "TenCate Grass",
    "97a76229c31eab35fa8215f023e65cc65277e2cf58a9663a7b3c4c2d7062cecc": "TigerTurf",
    "3b7cb9fafeaa72acc5d23736673069bdc1a295a2972db93e9e59914c306b2693": "GEO Surfaces",
}

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "rv": "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata",
    "rvrel": "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel",
}


class TemplateMismatch(Exception):
    """Raised when a fixed anchor cell doesn't contain the expected label."""


# --------------------------------------------------------------------------
# Rich-value ("picture in cell") resolution
# --------------------------------------------------------------------------
# Newer Excel can store an image as a cell's *value* (Insert > Pictures >
# Place in Cell). openpyxl has no support for this, so cells like Brand
# resolve to the literal string "#VALUE!". Resolving the real value requires
# manually walking the OOXML rich-data layer:
#   cell vm="N"  -->  xl/metadata.xml valueMetadata[N-1] --> rc/@v = i
#   xl/richData/rdrichvalue.xml  rv[i]  -->  first <v> = local image id
#   xl/richData/richValueRel.xml  rel[local image id]  -->  r:id
#   xl/richData/_rels/richValueRel.xml.rels  -->  r:id -> media/imageX.png
def _sheet_target_for(zf, sheet_name):
    wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {
        r.get("Id"): r.get("Target")
        for r in rels_xml.findall("rel:Relationship", NS)
    }
    for sheet in wb_xml.findall("main:sheets/main:sheet", NS):
        if sheet.get("name") == sheet_name:
            rid = sheet.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            target = rid_to_target[rid]
            return "xl/" + target if not target.startswith("xl/") else target
    raise TemplateMismatch(f"Sheet '{sheet_name}' not found in workbook.xml")


def build_rich_value_image_resolver(xlsx_path):
    """Returns a dict {cell_ref: sha256_hex} for every rich-value cell on
    SHEET_NAME, or {} if the workbook has no rich-data layer at all."""
    resolver = {}
    with zipfile.ZipFile(xlsx_path) as zf:
        names = set(zf.namelist())
        if "xl/richData/rdrichvalue.xml" not in names:
            return resolver

        sheet_path = _sheet_target_for(zf, SHEET_NAME)

        # cells with a vm= attribute on the target sheet
        sheet_xml = zf.read(sheet_path)
        vm_cells = re.findall(rb'<c r="([A-Z]+\d+)"[^>]*\bvm="(\d+)"', sheet_xml)
        if not vm_cells:
            return resolver

        metadata_xml = ET.fromstring(zf.read("xl/metadata.xml"))
        value_metadata = metadata_xml.findall("main:valueMetadata/main:bk", NS)

        rv_xml = ET.fromstring(zf.read("xl/richData/rdrichvalue.xml"))
        rv_entries = rv_xml.findall("rv:rv", NS)

        rel_xml = ET.fromstring(zf.read("xl/richData/richValueRel.xml"))
        rel_ids = [
            r.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            for r in rel_xml.findall("rvrel:rel", NS)
        ]

        rels_xml = ET.fromstring(zf.read("xl/richData/_rels/richValueRel.xml.rels"))
        rid_to_media = {
            r.get("Id"): r.get("Target")
            for r in rels_xml.findall("rel:Relationship", NS)
        }

        for cell_ref, vm in vm_cells:
            cell_ref = cell_ref.decode()
            vm_index = int(vm) - 1  # vm is 1-indexed
            if vm_index >= len(value_metadata):
                continue
            rc = value_metadata[vm_index].find("main:rc", NS)
            rv_index = int(rc.get("v"))
            if rv_index >= len(rv_entries):
                continue
            first_v = rv_entries[rv_index].find("rv:v", NS)
            local_image_id = int(first_v.text)
            if local_image_id >= len(rel_ids):
                continue
            rid = rel_ids[local_image_id]
            media_target = rid_to_media.get(rid)
            if not media_target:
                continue
            media_path = "xl/richData/" + media_target
            media_path = media_path.replace("xl/richData/../media/", "xl/media/")
            image_bytes = zf.read(media_path)
            resolver[cell_ref] = hashlib.sha256(image_bytes).hexdigest()
    return resolver


def resolve_brand_logo(image_hashes, cell_ref, warnings):
    digest = image_hashes.get(cell_ref)
    if digest is None:
        return None
    brand = KNOWN_BRAND_LOGOS.get(digest)
    if brand is None:
        warnings.append(
            f"Brand logo at {cell_ref} has unrecognized image hash {digest[:12]}... "
            "- add it to KNOWN_BRAND_LOGOS."
        )
        return f"UNKNOWN_LOGO:{digest[:12]}"
    return brand


# --------------------------------------------------------------------------
# Color code legend (Drop Box sheet: full name <-> abbreviation)
# --------------------------------------------------------------------------
def load_color_legend(wb):
    legend = {}
    if DROPBOX_SHEET_NAME not in wb.sheetnames:
        return legend
    ws = wb[DROPBOX_SHEET_NAME]
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=6).value  # F
        code = ws.cell(row=row, column=7).value  # G
        if name and code:
            legend[str(code).strip()] = str(name).strip()
    return legend


def color_name(legend, code):
    if code is None:
        return None
    return legend.get(str(code).strip(), str(code).strip())


# --------------------------------------------------------------------------
# Small helpers
# --------------------------------------------------------------------------
def require(ws, cell_ref, expected, warnings, label):
    actual = ws[cell_ref].value
    if actual != expected:
        warnings.append(
            f"Anchor mismatch: expected {cell_ref} == {expected!r} for {label}, "
            f"got {actual!r}. Template may have shifted - verify manually."
        )


def to_clean(value):
    if isinstance(value, str):
        value = value.strip()
    return value


# --------------------------------------------------------------------------
# Block 1: General Information  (B2:K10)
# --------------------------------------------------------------------------
def extract_general_information(ws, warnings):
    require(ws, "B2", "GENERAL INFORMATION", warnings, "general information header")
    return {
        "project_name": to_clean(ws["E3"].value),
        "customer": to_clean(ws["E5"].value),
        "customer_if_other": to_clean(ws["E6"].value),
        "purchase_order_number": to_clean(ws["E7"].value),
        "ship_pickup_date": ws["E8"].value.isoformat() if hasattr(ws["E8"].value, "isoformat") else to_clean(ws["E8"].value),
        "logo_fabricator": to_clean(ws["E9"].value),
        "pre_shipment_testing": to_clean(ws["E10"].value),
    }


# --------------------------------------------------------------------------
# Block 2: Product Specifications  (M2:U12)  + Block 3: MFG summary (M15:T16)
# --------------------------------------------------------------------------
def extract_product_specifications(ws, image_hashes, warnings):
    require(ws, "M2", "PRODUCT SPECIFICATIONS", warnings, "product specifications header")

    def num_or_none(v):
        return v if isinstance(v, (int, float)) else None

    yarns = []
    for label_cell, type_cell, ozsy_cell in (("M6", "O6", "Q6"), ("M7", "O7", "Q7"), ("M8", "O8", "Q8"), ("M9", "O9", "Q9")):
        yarn_type = ws[type_cell].value
        if yarn_type in (None, 0, "0"):
            continue  # unused creel/yarn slot
        yarns.append({
            "slot": ws[label_cell].value,
            "yarn_type": to_clean(yarn_type),
            "face_weight_contribution_oz_sy": num_or_none(ws[ozsy_cell].value),
        })

    return {
        "product_name": to_clean(ws["M4"].value),
        "brand": resolve_brand_logo(image_hashes, "M5", warnings),
        "finished_pile_height_in": num_or_none(ws["R4"].value),
        "tufted_pile_height_in": num_or_none(ws["U5"].value),
        "face_weight_oz": num_or_none(ws["T4"].value),
        "primary_tufting_gauge_in": num_or_none(ws["U6"].value),
        "secondary_tufting_gauge_in": num_or_none(ws["U7"].value) if ws["U7"].value != "-" else None,
        "thread_up": to_clean(ws["T8"].value),
        "stitch_rate": to_clean(ws["T9"].value),
        "yarns": yarns,
        "primary_backing": {
            "type": to_clean(ws["O10"].value),
            "weight_oz_sy": num_or_none(ws["Q10"].value),
        },
        "secondary_coating": {
            "type": to_clean(ws["O11"].value),
            "weight_oz_sy": num_or_none(ws["Q11"].value),
        },
        "product_type": to_clean(ws["T10"].value),
        "perforations": to_clean(ws["T11"].value) if ws["U11"].value == "-" else to_clean(ws["U11"].value),
        "roll_width_in": num_or_none(ws["T12"].value),
        "product_total_weight_oz_sy": num_or_none(ws["Q12"].value),
    }


def extract_mfg_summary(ws, warnings):
    require(ws, "M15", "MFG Rolls", warnings, "MFG summary header")
    return {
        "mfg_rolls": ws["M16"].value,
        "mfg_lf": ws["N16"].value,
        "mfg_sf": ws["O16"].value,
        "mfg_weight_lbs": ws["P16"].value,
        "ltl_trucks_10k": ws["R16"].value,
        "ftl_trucks_40k": ws["T16"].value,
    }


# --------------------------------------------------------------------------
# Block 4: Yarn SKUs  (rows 671-676, columns A-AB)
# --------------------------------------------------------------------------
def extract_yarn_skus(ws, legend, warnings):
    require(ws, "A672", "Creel Position", warnings, "yarn SKU header")

    color_columns = []  # (col_index, code)
    for col in range(7, 29):  # G..AB
        code = ws.cell(row=672, column=col).value
        if code:
            color_columns.append((col, str(code).strip()))

    creels = []
    for row in range(673, 677):
        yarn_type = ws.cell(row=row, column=4).value  # D
        if yarn_type in (None, 0, "0"):
            continue
        colors = []
        for col, code in color_columns:
            sku = ws.cell(row=row, column=col).value
            if sku is not None:
                colors.append({
                    "color_code": code,
                    "color_name": color_name(legend, code),
                    "sku": sku,
                })
        creels.append({
            "creel_position": ws.cell(row=row, column=1).value,      # A
            "creel_marker_color": ws.cell(row=row, column=2).value,  # B
            "yarn_type": to_clean(yarn_type),
            "available_colors": colors,
        })
    return creels


# --------------------------------------------------------------------------
# Block 5: Roll layout table (row 684 onward, ends at literal "Creel Change")
# --------------------------------------------------------------------------
SEGMENT_COLUMNS = [(10, 11), (12, 13), (14, 15), (16, 17), (18, 19),
                   (20, 21), (22, 23), (24, 25), (26, 27), (28, 29)]  # J/K .. AB


def _find_table_end(ws, start_row, max_scan=1000):
    row = start_row
    while row < start_row + max_scan:
        if ws.cell(row=row, column=2).value == "Creel Change":
            return row
        row += 1
    return None


def layout_signature(segments):
    """Identity of a roll's physical layout: the ordered width/colour
    segments across the roll, left to right. Length (mfg_roll_length_lf)
    plays no part - two rolls with this same signature are threaded
    identically and are interchangeable for sequencing purposes."""
    return tuple((s["width_in"], s["color_code"]) for s in segments)


def extract_rolls(ws, legend, warnings):
    require(ws, "A680", "Sort", warnings, "roll table header")

    start_row = 684
    end_row = _find_table_end(ws, start_row)
    if end_row is None:
        warnings.append("Could not find 'Creel Change' terminator; roll table may be truncated.")
        end_row = ws.max_row + 1

    rolls = []
    setup_group = 1
    current_roll = None

    for row in range(start_row, end_row):
        marker = ws.cell(row=row, column=2).value
        if marker == "SETUP CHANGE":
            setup_group += 1
            current_roll = None
            continue

        sort_val = ws.cell(row=row, column=1).value
        lot = ws.cell(row=row, column=2).value
        segments = []
        for wcol, ccol in SEGMENT_COLUMNS:
            width = ws.cell(row=row, column=wcol).value
            code = ws.cell(row=row, column=ccol).value
            if width is None or code is None:
                continue
            segments.append({
                "width_in": width,
                "color_code": str(code).strip(),
                "color_name": color_name(legend, code),
            })

        is_blank_padding = sort_val is None and lot is None and not segments
        if is_blank_padding:
            continue

        if lot is not None:
            # a new roll-defining row
            width_total = sum(s["width_in"] for s in segments)
            roll_width = ws.cell(row=row, column=9).value  # I
            if roll_width is not None and abs(width_total - roll_width) > 0.01:
                warnings.append(
                    f"Row {row}: segment widths sum to {width_total}, "
                    f"expected roll width {roll_width}."
                )
            current_roll = {
                "row": row,
                "setup_group": setup_group,
                "sort": sort_val,
                "navision_lot": to_clean(lot),
                "panel_numbers": to_clean(ws.cell(row=row, column=4).value),  # D
                "roll_type": to_clean(ws.cell(row=row, column=5).value),     # E
                "roll_qty": ws.cell(row=row, column=6).value,                # F
                "mfg_roll_length_lf": ws.cell(row=row, column=7).value,      # G
                "total_mfg_sf": ws.cell(row=row, column=8).value,            # H
                "roll_width_in": roll_width,
                "segments": segments,
                "additional_panel_layouts": [],
                "layout_signature": layout_signature(segments),
            }
            rolls.append(current_roll)
        else:
            # continuation row: extra panel cut from the previous roll, same setup group
            if current_roll is None:
                warnings.append(
                    f"Row {row}: continuation row with no preceding roll in this "
                    "setup group - skipped."
                )
                continue
            width_total = sum(s["width_in"] for s in segments) if segments else None
            if width_total is not None and current_roll["roll_width_in"] is not None \
                    and abs(width_total - current_roll["roll_width_in"]) > 0.01:
                warnings.append(
                    f"Row {row}: continuation segment widths sum to {width_total}, "
                    f"expected roll width {current_roll['roll_width_in']}."
                )
            current_roll["additional_panel_layouts"].append({
                "row": row,
                "roll_type": to_clean(ws.cell(row=row, column=5).value),
                "segments": segments,
            })

    group_ids = {}
    for roll in rolls:
        sig = roll["layout_signature"]
        if sig not in group_ids:
            group_ids[sig] = len(group_ids) + 1
        roll["layout_group"] = group_ids[sig]
        roll["layout_signature"] = "|".join(f"{w}{c}" for w, c in sig)

    return rolls, setup_group


# --------------------------------------------------------------------------
# Top-level extraction
# --------------------------------------------------------------------------
def extract_workbook(xlsx_path):
    warnings = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise TemplateMismatch(f"Workbook has no '{SHEET_NAME}' sheet")
    ws = wb[SHEET_NAME]

    legend = load_color_legend(wb)
    image_hashes = build_rich_value_image_resolver(xlsx_path)

    general_information = extract_general_information(ws, warnings)
    product_specifications = extract_product_specifications(ws, image_hashes, warnings)
    mfg_summary = extract_mfg_summary(ws, warnings)
    yarn_skus = extract_yarn_skus(ws, legend, warnings)
    rolls, num_setup_groups = extract_rolls(ws, legend, warnings)

    return {
        "source_file": Path(xlsx_path).name,
        "general_information": general_information,
        "product_specifications": product_specifications,
        "mfg_summary": mfg_summary,
        "yarn_skus": yarn_skus,
        "rolls": rolls,
        "roll_count": len(rolls),
        "distinct_layout_count": len({r["layout_group"] for r in rolls}),
        "setup_change_count": num_setup_groups - 1,
        "warnings": warnings,
    }


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("files", nargs="+", help="Path(s) to FIELD LAYOUT .xlsx workbooks")
    parser.add_argument("-o", "--out-dir", default=".", help="Directory to write JSON output into")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    exit_code = 0
    for file_path in args.files:
        try:
            result = extract_workbook(file_path)
        except TemplateMismatch as exc:
            print(f"[FAIL] {file_path}: {exc}", file=sys.stderr)
            exit_code = 1
            continue

        out_path = out_dir / (Path(file_path).stem + ".json")
        out_path.write_text(json.dumps(result, indent=2, default=str))

        status = "OK" if not result["warnings"] else f"OK with {len(result['warnings'])} warning(s)"
        print(f"[{status}] {file_path} -> {out_path} ({result['roll_count']} rolls)")
        for w in result["warnings"]:
            print(f"    warning: {w}")

    sys.exit(exit_code)


if __name__ == "__main__":
    main()
